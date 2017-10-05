#!/usr/bin/env python

'''
Purpose:    Connect to routers/firewalls and record ZScaler config info
Author:
            ___  ____ _ ____ _  _    _  _ _    ____ ___ ___
            |__] |__/ | |__| |\ |    |_/  |    |  |  |    /
            |__] |  \ | |  | | \|    | \_ |___ |__|  |   /__
            Brian.Klotz@nike.com

Version:    1.0
Date:       October 2017
'''
import argparse
import netmiko
import getpass
import logging
import os
import openpyxl
import datetime

# Set up argument parser and help info
parser = argparse.ArgumentParser(description='Connect to list of devices and \
                                 run a set of commands on each to get Zscaler \
                                 tunnel information')
always_required = parser.add_argument_group('always required')
always_required.add_argument("devices", nargs=1,
                             help="Name of file containing devices",
                             metavar='<devices_file>')
args = parser.parse_args()

# Configure logging
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
handler = logging.FileHandler('output.log')
handler.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s - %(message)s')
handler.setFormatter(formatter)
logger.addHandler(handler)


def open_file(file):
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    input_info = {}
    for row in range(2, ws.max_row + 1):
        if ws['A' + str(row)].value:  # Prevents adding blank lines
            device = ws['A' + str(row)].value
            # Subtract 1 from the row so that devices are numbered 1, 2, 3...
            input_info[row - 1] = {'host': device,
                                   'device_type': ws['B' + str(row)].value,
                                   }
    return input_info


def get_creds():  # Prompt for credentials
    username = getpass.getuser()
#   username = raw_input('User ID: ')
    try:
        password = getpass.getpass()
        return username, password
    except KeyboardInterrupt:
        print('\n')
        exit()


def indentem(lines):
    lines = ' '*26 + lines
    lines = lines.replace('\n', '\n' + ' '*26)
    return lines


def send_and_log(logger, connection, command):
    logger.info('Sending command:\n' + ' '*26 + command.strip())
    output = connection.send_command(command)
    indented_output = indentem(output.strip())
    logger.info('Output:\n' + indented_output)
    return output


def main():
    device_file = args.devices[0]
    input_info = open_file(device_file)

    username, password = get_creds()

    netmiko_exceptions = (netmiko.ssh_exception.NetMikoTimeoutException,
                          netmiko.ssh_exception.NetMikoAuthenticationException)

    # Prep the output file
    today = datetime.date.today()
    filename = 'Zscaler_info_' + str(today) + '.xlsx'
    filename = os.path.join(os.environ['TMPDIR'], filename)
    wb = openpyxl.Workbook()
    wb.save(filename=filename)
    ws1 = wb.active

    # Prep ASA output
    ws1.title = 'ASA Zscaler Node Info'
    ws1['A1'] = 'Hostname'
    ws1['B1'] = 'Zscaler_Node_1'
    ws1['C1'] = 'Zscaler_Node_2'
    ws1['D1'] = 'Source Interface'
    ws1['E1'] = 'Source IP'

    # Prep Router output
    ws2 = wb.create_sheet(title='Router Zscaler Node Info')
    ws2['A1'] = 'Hostname'
    ws2['B1'] = 'Zscaler_Node_1'
    ws2['C1'] = 'Zscaler_Node_1_NextHop'
    ws2['D1'] = 'Zscaler_Node_1_Source_Interface'
    ws2['E1'] = 'Zscaler_Node_1_Source_IP'
    ws2['F1'] = 'Zscaler_Node_2'
    ws2['G1'] = 'Zscaler_Node_2_NextHop'
    ws2['H1'] = 'Zscaler_Node_2_Source_Interface'
    ws2['I1'] = 'Zscaler_Node_2_Source_IP'

    asa_index = 1  # Start count at 1 so entries start at row 2
    rtr_index = 1
    asa_peers_command = 'show run crypto map | i 65000 set peer'
    asa_crypto_command = 'show run crypto map | i interface'

    start = datetime.datetime.now()
    print('Start time: ' + str(start))

    # Build dictionary of devices
    for row in range(1, len(input_info) + 1):
        device_dict = {'host': input_info[row]['host'],
                       'device_type': input_info[row]['device_type'],
                       'username': username,
                       'password': password,
                       'secret': password
                       }
        asa_peers_output = ''
        zscaler_node_1 = ''
        zscaler_node_2 = ''
        crypto_output = ''
        iface = ''
        sh_int_output = ''
        source_IP = ''
        sh_int_tun = ''

        print('-'*79)
        print('Connecting to ' + device_dict['host'] + ' (' +
              str(row) + ' of ' + str(len(input_info)) + ')')

        try:  # Connect to device
            connection = netmiko.ConnectHandler(**device_dict)
            logger.info('-'*25 + '\n' + ' '*26 + 'Successfully connected to %s',
                        device_dict['host'])
            connection.enable()
            print('Gathering data and writing to output file...')

            # Gather information
            if device_dict['device_type'] == 'cisco_asa':
                asa_index += 1
                ws1['A' + str(asa_index)] = device_dict['host']
                asa_peers_output = send_and_log(
                    logger, connection, asa_peers_command)
                if asa_peers_output:
                    zscaler_node_1 = asa_peers_output.split()[-2]
                    zscaler_node_2 = asa_peers_output.split()[-1]
                    crypto_output = send_and_log(
                        logger, connection, asa_crypto_command)
                    iface = crypto_output.split()[-1]
                    cmd = 'show int ' + iface + ' | inc IP'
                    sh_int_output = send_and_log(
                        logger, connection, cmd
                        )
                    source_IP = sh_int_output.split()[2]
                    source_IP = source_IP.strip(',')

                    ws1['B' + str(asa_index)] = zscaler_node_1
                    ws1['C' + str(asa_index)] = zscaler_node_2
                    ws1['D' + str(asa_index)] = iface
                    ws1['E' + str(asa_index)] = source_IP
                else:
                    ws1['B' + str(asa_index)] = 'No peers found!'

            if device_dict['device_type'] == 'cisco_ios':
                rtr_index += 1
                ws2['A' + str(rtr_index)] = device_dict['host']
                for tunnel in ['tun1028', 'tun1128']:
                    cmd = 'sh ip int ' + tunnel + ' | in protocol'
                    tunnels_up = send_and_log(
                        logger, connection, cmd
                        )
                    if 'line protocol' in tunnels_up:
                        zscaler_node = ''
                        tun_gw_output = ''
                        zscaler_node_nexthop = ''
                        tun_src_int_output = ''
                        tun_src_ip_output = ''
                        tun_src_ip = ''
                    # Get tunnel destination IP
                        cmd = 'sh run int ' + tunnel + ' | in destination'
                        sh_int_tun = send_and_log(
                            logger, connection, cmd
                            )
                        zscaler_node = sh_int_tun.split()[-1]

                    # Get next hop IP
                        cmd = 'sh run | i route ' + zscaler_node
                        tun_gw_output = send_and_log(
                            logger, connection, cmd
                            )
                        if tun_gw_output:
                            zscaler_node_nexthop = tun_gw_output.split()[4]
                        else:
                            zscaler_node_nexthop = 'Not configured'

                    # Get tunnel source interface
                        cmd = 'sh run int ' + tunnel + ' | i source'
                        tun_src_int_output = send_and_log(
                            logger, connection, cmd
                            )
                        tun_src = tun_src_int_output.split()[2]

                    # Get Tunnel source IP
                        cmd = 'sh ip int ' + tun_src + ' | i Internet address'
                        tun_src_ip_output = send_and_log(
                            logger, connection, cmd
                            )
                        tun_src_ip = tun_src_ip_output.split()[3]
                        tun_src_ip = tun_src_ip.split('/')[0]

                    # Write output to file
                        if tunnel == 'tun1028':
                            ws2['B' + str(rtr_index)] = zscaler_node
                            ws2['C' + str(rtr_index)] = zscaler_node_nexthop
                            ws2['D' + str(rtr_index)] = tun_src
                            ws2['E' + str(rtr_index)] = tun_src_ip
                        elif tunnel == 'tun1128':
                            ws2['F' + str(rtr_index)] = zscaler_node
                            ws2['G' + str(rtr_index)] = zscaler_node_nexthop
                            ws2['H' + str(rtr_index)] = tun_src
                            ws2['I' + str(rtr_index)] = tun_src_ip
                    else:
                        ws2['B' + str(rtr_index)] = 'Tunnels not configured'
            # Disconnect from device
            print('Disconnecting...')
            logger.info('Disconnecting from %s' % device_dict['host'])
            connection.disconnect()
            wb.save(filename)

        except netmiko_exceptions as e:
            print('Failed to connect: %s' % e)
            logger.error('Failed to connect %s', e)

    end = datetime.datetime.now()
    print('End time: ' + str(end))
    elapsed = end - start
    print('Elapsed time: ' + str(elapsed))
    #  Format output columns
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 13
    ws1.column_dimensions['C'].width = 13
    ws1.column_dimensions['D'].width = 14
    ws1.column_dimensions['E'].width = 14
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 13
    ws2.column_dimensions['C'].width = 21
    ws2.column_dimensions['D'].width = 26
    ws2.column_dimensions['E'].width = 22
    ws2.column_dimensions['F'].width = 13
    ws2.column_dimensions['G'].width = 21
    ws2.column_dimensions['H'].width = 27
    ws2.column_dimensions['I'].width = 21
    wb.save(filename)
    # Open spreadsheet
    #os.system('open %s' % filename)


main()
